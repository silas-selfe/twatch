"""Export traffic.db to an Excel workbook.

Sheets:
  About     -- provenance: source DB, export time, runs, what each sheet means
  Events    -- one row per counted road user (raw data, local time)
  Hourly    -- per-hour counts by class + direction via COUNTIFS over Events,
               plus measured uptime %% from heartbeats
  Daily     -- per-day rollup, formulas over Events/Hourly

Counts are live COUNTIFS formulas, so the summaries stay correct if rows are
ever corrected or filtered. Uptime is measured source data (heartbeat rows),
not a formula, and is labeled as such.

Usage: .venv/bin/python export_xlsx.py [--out traffic.xlsx]
"""
from __future__ import annotations

import argparse
import sqlite3
from datetime import datetime, timezone
from pathlib import Path

import yaml
from openpyxl import Workbook
from openpyxl.comments import Comment
from openpyxl.styles import Font
from openpyxl.utils import get_column_letter

HERE = Path(__file__).resolve().parent
ARIAL = Font(name="Arial", size=10)
ARIAL_BOLD = Font(name="Arial", size=10, bold=True)


def style_row(ws, row, bold=False):
    for cell in ws[row]:
        cell.font = ARIAL_BOLD if bold else ARIAL


def local(ts_utc: str) -> datetime:
    return datetime.fromisoformat(ts_utc).astimezone().replace(tzinfo=None)


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--out", default=str(HERE / "traffic.xlsx"))
    args = ap.parse_args()

    from watch import load_config
    cfg = load_config(str(HERE / "config.yaml"))
    hb_secs = cfg["output"]["heartbeat_seconds"]
    con = sqlite3.connect(str(HERE / cfg["output"]["db"]))

    events = con.execute(
        "SELECT ts_utc, class, direction, conf_mean, frames_seen, duration_s,"
        " speed_px_s FROM events ORDER BY ts_utc").fetchall()
    if not events:
        raise SystemExit("no events in database")
    classes = [r[0] for r in con.execute(
        "SELECT DISTINCT class FROM events ORDER BY class")]
    directions = [r[0] for r in con.execute(
        "SELECT DISTINCT direction FROM events ORDER BY direction")]
    runs = con.execute(
        "SELECT id, started_utc, ended_utc, model FROM runs ORDER BY id").fetchall()

    # measured uptime: heartbeat coverage per local hour
    hb_per_hour: dict[datetime, int] = {}
    for (ts,) in con.execute("SELECT ts_utc FROM heartbeats"):
        h = local(ts).replace(minute=0, second=0, microsecond=0)
        hb_per_hour[h] = hb_per_hour.get(h, 0) + 1

    wb = Workbook()

    # ---------------- Events ----------------
    ev = wb.active
    ev.title = "Events"
    headers = ["Timestamp (local)", "Date", "Hour", "Class", "Direction",
               "Confidence", "Frames Seen", "Duration (s)", "Speed (px/s)"]
    ev.append(headers)
    style_row(ev, 1, bold=True)
    for i, (ts, cls, direction, conf, frames, dur, speed) in enumerate(events):
        r = i + 2
        ev.append([local(ts), f"=INT(A{r})", f"=HOUR(A{r})", cls, direction,
                   conf, frames, dur, speed])
    n = len(events) + 1  # last data row
    for row in ev.iter_rows(min_row=2):
        for cell in row:
            cell.font = ARIAL
        row[0].number_format = "yyyy-mm-dd hh:mm:ss"
        row[1].number_format = "yyyy-mm-dd"
        row[2].number_format = "0"
        row[5].number_format = "0.00"
        row[7].number_format = "0.0"
        row[8].number_format = "0"
    widths = [19, 12, 6, 11, 14, 11, 11, 12, 12]
    for i, w in enumerate(widths, 1):
        ev.column_dimensions[get_column_letter(i)].width = w
    ev.freeze_panes = "A2"

    # ---------------- Hourly ----------------
    first_h = local(events[0][0]).replace(minute=0, second=0, microsecond=0)
    last_h = local(events[-1][0]).replace(minute=0, second=0, microsecond=0)
    hours = []
    h = first_h
    from datetime import timedelta
    while h <= last_h:
        hours.append(h)
        h += timedelta(hours=1)

    hr = wb.create_sheet("Hourly")
    hr.append(["Hour", "Uptime %"] + [c.capitalize() for c in classes]
              + ["Total"] + [d for d in directions])
    style_row(hr, 1, bold=True)
    hr["B1"].comment = Comment(
        "Measured, not computed: heartbeat rows written every "
        f"{hb_secs}s while the collector ran. Hours below 100% include "
        "downtime -- their counts are undercounts of reality.", "trafficwatch")
    ts_col, cls_col, dir_col = "A", "D", "E"
    for i, h in enumerate(hours):
        r = i + 2
        uptime = min(hb_per_hour.get(h, 0) * hb_secs / 3600, 1.0)
        lo = f"Events!${ts_col}$2:${ts_col}${n}"
        base = (f'{lo},">="&$A{r},{lo},"<"&($A{r}+TIME(1,0,0))')
        row = [h, uptime]
        for j, c in enumerate(classes):
            col = get_column_letter(3 + j)
            row.append(f'=COUNTIFS({base},Events!${cls_col}$2:${cls_col}${n},{col}$1)')
        first_c, last_c = get_column_letter(3), get_column_letter(2 + len(classes))
        row.append(f"=SUM({first_c}{r}:{last_c}{r})")
        for j, d in enumerate(directions):
            col = get_column_letter(4 + len(classes) + j)
            row.append(f'=COUNTIFS({base},Events!${dir_col}$2:${dir_col}${n},{col}$1)')
        hr.append(row)
    # headers must be the raw class strings: COUNTIFS matches against them
    for j, c in enumerate(classes):
        hr.cell(row=1, column=3 + j, value=c)
    for row in hr.iter_rows(min_row=2):
        for cell in row:
            cell.font = ARIAL
        row[0].number_format = "yyyy-mm-dd hh:mm"
        row[1].number_format = "0.0%"
    hr.column_dimensions["A"].width = 17
    for i in range(2, 5 + len(classes) + len(directions)):
        hr.column_dimensions[get_column_letter(i)].width = 13
    hr.freeze_panes = "A2"
    m = len(hours) + 1  # last hourly row

    # ---------------- Daily ----------------
    days = sorted({h.date() for h in hours})
    dy = wb.create_sheet("Daily")
    dy.append(["Date"] + [c for c in classes] + ["Total", "Mean Uptime %"])
    style_row(dy, 1, bold=True)
    dy.cell(row=1, column=len(classes) + 3).comment = Comment(
        "Average of hourly uptime across the 24 hours of the date. A low "
        "value means the day's total is a partial count.", "trafficwatch")
    for i, d in enumerate(days):
        r = i + 2
        lo = f"Events!${ts_col}$2:${ts_col}${n}"
        base = f'{lo},">="&$A{r},{lo},"<"&($A{r}+1)'
        row = [datetime(d.year, d.month, d.day)]
        for j, c in enumerate(classes):
            col = get_column_letter(2 + j)
            row.append(f'=COUNTIFS({base},Events!${cls_col}$2:${cls_col}${n},{col}$1)')
        first_c, last_c = get_column_letter(2), get_column_letter(1 + len(classes))
        row.append(f"=SUM({first_c}{r}:{last_c}{r})")
        row.append(f'=AVERAGEIFS(Hourly!$B$2:$B${m},Hourly!$A$2:$A${m},'
                   f'">="&$A{r},Hourly!$A$2:$A${m},"<"&($A{r}+1))')
        dy.append(row)
    for row in dy.iter_rows(min_row=2):
        for cell in row:
            cell.font = ARIAL
        row[0].number_format = "yyyy-mm-dd"
        row[-1].number_format = "0.0%"
    dy.column_dimensions["A"].width = 12
    for i in range(2, len(classes) + 4):
        dy.column_dimensions[get_column_letter(i)].width = 13
    dy.freeze_panes = "A2"

    # ---------------- About ----------------
    ab = wb.create_sheet("About", 0)
    tz = datetime.now().astimezone().tzname()
    lines = [
        ("trafficwatch data export", True),
        ("", False),
        (f"Source: {HERE / cfg['output']['db']}", False),
        (f"Exported: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')} "
         f"(all timestamps in local time, {tz})", False),
        (f"Events: {len(events)} counted road users, "
         f"{events[0][0][:10]} to {events[-1][0][:10]} (UTC dates)", False),
        ("", False),
        ("Sheets", True),
        ("Events -- one row per counted road user (raw data).", False),
        ("Hourly -- counts by class and direction; formulas over Events."
         " Uptime % is measured from collector heartbeats.", False),
        ("Daily -- per-day rollup; formulas over Events and Hourly.", False),
        ("", False),
        ("Reading the numbers honestly", True),
        ("An hour with uptime below 100% has counts LOWER than real traffic"
         " because the collector was not running the whole hour."
         " Filter to uptime = 100% for unbiased comparisons.", False),
        ("", False),
        ("Collector runs in this export", True),
        ("run id | started (local) | ended (local) | model", True),
    ]
    for i, (text, bold) in enumerate(lines, 1):
        c = ab.cell(row=i, column=1, value=text)
        c.font = ARIAL_BOLD if bold else ARIAL
    r0 = len(lines) + 1
    for i, (rid, s, e, model) in enumerate(runs):
        vals = [rid, local(s).strftime("%Y-%m-%d %H:%M"),
                local(e).strftime("%Y-%m-%d %H:%M") if e else "(running)", model]
        for j, v in enumerate(vals):
            c = ab.cell(row=r0 + i, column=1 + j, value=v)
            c.font = ARIAL
    ab.column_dimensions["A"].width = 100

    wb.save(args.out)
    print(f"wrote {args.out}: {len(events)} events, {len(hours)} hours, "
          f"{len(days)} days")


if __name__ == "__main__":
    main()
